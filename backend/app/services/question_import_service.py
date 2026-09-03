import csv
import io
import re
from typing import Optional

from fastapi import HTTPException, UploadFile, status
import openpyxl
from pypdf import PdfReader

from app.services.module_authoring_service import DEFAULT_PART_HEADINGS



MAX_PDF_BYTES = 15 * 1024 * 1024
MAX_CSV_BYTES = 5 * 1024 * 1024
MAX_EXCEL_BYTES = 20 * 1024 * 1024
MAX_EXTRACTED_TEXT = 250_000
MAX_IMPORT_QUESTIONS = 500

QUESTION_RE = re.compile(r"^(?:q(?:uestion)?\s*)?(\d{1,4})\s*[.)\-:]\s*(.*)$", re.IGNORECASE)
OPTION_RE = re.compile(r"^\(?([A-Z])\)?\s*[.)\-:]\s*(.+)$", re.IGNORECASE)
ANSWER_RE = re.compile(r"^(?:correct\s+)?answer(?:s)?\s*[:\-]\s*(.+)$", re.IGNORECASE)
EXPLANATION_RE = re.compile(r"^(?:explanation|rationale)\s*[:\-]\s*(.+)$", re.IGNORECASE)
PART_HEADER_RE = re.compile(r"^(?:part|section)\s*[:\-]\s*(.+)$", re.IGNORECASE)
ANSWER_KEY_HEADER_RE = re.compile(r"^(?:answers?|answer\s+key)\s*:?$", re.IGNORECASE)
ANSWER_KEY_ENTRY_RE = re.compile(
    r"^(\d{1,4})\s*[.)\-:]?\s*(?:answer\s*[:\-]\s*)?([A-Z](?:\s*(?:,|;|/|&|\band\b)\s*[A-Z])*)$",
    re.IGNORECASE,
)


def _clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _clean_multiline(value: object) -> str:
    """Like _clean, but keeps line breaks - a shared passage (or a notepad's
    heading-then-gapped-paragraph text) depends on them to tell its lines
    apart; collapsing to one line loses the heading and runs every gap
    together."""
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in str(value or "").splitlines()]
    return "\n".join(lines).strip()


def _split_answers(value: object) -> list[str]:
    text = _clean(value).upper()
    if not text:
        return []
    text = re.sub(r"^(?:OPTION|ANSWER)\s+", "", text)
    parts = re.split(r"\s*(?:,|;|\||/|\band\b)\s*", text)
    return list(dict.fromkeys(part.strip(" ().") for part in parts if part.strip(" ().")))


def _normalize_cloze_passage(text: str) -> str:
    if not text:
        return text
    # 1. (1) ______ or (1) ___ or (1) _
    text = re.sub(r"\(\s*(\d+)\s*\)\s*_{1,}", r"{{blank:\1}}", text)
    # 2. ___ (1) ___ or __(1)
    text = re.sub(r"_{1,}\s*\(\s*(\d+)\s*\)\s*_{0,}", r"{{blank:\1}}", text)
    # 3. [1] ___ or [1]
    text = re.sub(r"\[\s*(\d+)\s*\]\s*_{0,}", r"{{blank:\1}}", text)
    # 4. Standalone (1) if surrounded by spaces or punctuation: e.g. " ideas (1) . "
    text = re.sub(r"(?<=\s)\(\s*(\d+)\s*\)(?=[\s,.:;!?])", r"{{blank:\1}}", text)
    return text


def _infer_type(
    options: list[dict],
    answers: list[str],
    supplied: str = "",
    target_part: str = "",
) -> str:
    aliases = {
        "mcq": "mcq_single",
        "multiple_choice": "mcq_single",
        "multiple choice": "mcq_single",
        "single_choice": "mcq_single",
        "multi_select": "mcq_multiple",
        "multiple_answers": "mcq_multiple",
        "true_false": "true_false_not_given",
        "true false not given": "true_false_not_given",
        "yes no not given": "yes_no_not_given",
        "matching once": "matching_unique",
        "matching unique": "matching_unique",
        "matching reusable": "matching_reusable",
        "short answer": "short_answer",
        "fill in the blank": "fill_blank",
        "writing": "essay",
        "speaking": "speaking_prompt",
    }
    normalized = _clean(supplied).lower().replace("-", "_")
    if normalized in aliases:
        return aliases[normalized]
    valid = {
        "mcq_single", "mcq_multiple", "true_false_not_given", "yes_no_not_given",
        "short_answer", "fill_blank", "matching_unique", "matching_reusable",
        "essay", "speaking_prompt",
    }
    if normalized in valid:
        return normalized
    part_lower = _clean(target_part).lower()
    if "writing" in part_lower:
        return "essay"
    if "speaking" in part_lower:
        return "speaking_prompt"
    option_values = {_clean(option.get("text")).lower() for option in options}
    if {"true", "false"}.issubset(option_values):
        return "true_false_not_given"
    if {"yes", "no"}.issubset(option_values):
        return "yes_no_not_given"
    if options:
        return "mcq_multiple" if len(answers) > 1 else "mcq_single"
    return "short_answer"


DEFAULT_PART_HEADINGS: dict[str, str] = {
    "reading_1a": "Read each sentence. Choose the word that can best replace the bold word without changing the meaning.",
    "reading_1b": "Read the text and choose the correct word for each gap.",
    "reading_2": "Read the text. Six sentences have been removed. Choose the sentence that best fits each gap. One sentence is a distractor.",
    "reading_3": "Read texts A–D. For questions 18–24, decide which text answers the question.",
    "reading_4": "Read the text and choose the correct answer for each question.",
    "listening_1": "You will hear some short conversations. You will hear each conversation twice. Choose the correct answer to complete each conversation.",
    "listening_2": "You will hear five conversations. Listen to the conversations and answer the questions. Choose the correct answer. You will hear each conversation twice.",
    "listening_3": "You will hear a recording. You will hear the recording twice. Complete the notes with NO MORE THAN THREE WORDS for each gap.",
    "listening_4": "You will hear a discussion. You will hear the discussion twice. Choose the correct answer for each question.",
}


def _question_preview(
    *,
    prompt: str,
    options: list[dict],
    correct_answers: list[str],
    question_type: str = "",
    instructions: str = "",
    part_heading: str = "",
    passage: str = "",
    explanation: str = "",
    points: object = 1,
    difficulty: str = "medium",
    group_label: str = "",
    turn_type: str = "",
    preparation_seconds: object = None,
    response_seconds: object = None,
    adaptive_follow_up: object = False,
    target_part: str = "",
) -> dict:
    normalized_options = []
    for index, option in enumerate(options):
        key = _clean(option.get("key")).upper() or chr(65 + index)
        text = _clean(option.get("text"))
        if text:
            normalized_options.append({"key": key, "text": text})

    answers = _split_answers("|".join(correct_answers))
    option_keys = {option["key"] for option in normalized_options}
    option_text_to_key = {option["text"].upper(): option["key"] for option in normalized_options}
    answers = [option_text_to_key.get(answer, answer) for answer in answers]
    kind = _infer_type(normalized_options, answers, question_type, target_part=target_part)

    warnings: list[str] = []
    is_l1 = bool(
        target_part
        and (
            "listening_1" in target_part.lower()
            or "listening 1" in target_part.lower()
            or (target_part.lower().endswith("_1") and "listen" in target_part.lower())
        )
    )
    part_clean_lower = _clean(target_part).lower().replace(" ", "_")
    is_cloze_gap = "reading_1b" in part_clean_lower or "reading_2" in part_clean_lower
    if not _clean(prompt):
        if is_l1:
            prompt = "Question"
        elif is_cloze_gap:
            prompt = "Gap"
        else:
            warnings.append("Question text is missing")
    if kind in {"mcq_single", "mcq_multiple", "matching_unique", "matching_reusable"} and len(normalized_options) < 2:
        warnings.append("At least two choices are required")
    part_lower = _clean(target_part).lower()
    is_subjective = kind in {"essay", "speaking_prompt"} or "writing" in part_lower or "speaking" in part_lower
    is_speaking_three = "speaking_3" in part_lower.replace(" ", "_") or "speaking 3" in part_lower
    if is_speaking_three and kind == "speaking_prompt" and not str(passage or "").strip():
        read_aloud_match = re.match(
            r"^\s*read\s+(?:the\s+)?(?:short\s+|given\s+)?text\s+aloud[\s.:-]+\s*(.+)$",
            prompt,
            re.IGNORECASE | re.DOTALL,
        )
        if read_aloud_match and read_aloud_match.group(1).strip():
            prompt = "Read the given text aloud."
            passage = read_aloud_match.group(1).strip()
    needs_answer = not is_subjective
    if needs_answer and not answers:
        warnings.append("Correct answer was not detected")
    if normalized_options and any(answer not in option_keys for answer in answers):
        warnings.append("A detected answer does not match an option key")
    try:
        numeric_points = float(points or 1)
        if numeric_points <= 0:
            raise ValueError
    except (TypeError, ValueError):
        numeric_points = 1
        warnings.append("Invalid points value was replaced with 1")
    difficulty = _clean(difficulty).lower()
    if difficulty not in {"easy", "medium", "hard"}:
        difficulty = "medium"
        warnings.append("Invalid difficulty was replaced with medium")

    interaction: dict[str, object] = {}
    if _clean(group_label):
        interaction["group_label"] = _clean(group_label)
    if _clean(turn_type):
        interaction["turn_type"] = _clean(turn_type).lower().replace(" ", "_")
    for field, raw in (("preparation_seconds", preparation_seconds), ("response_seconds", response_seconds)):
        if _clean(raw):
            try:
                interaction[field] = int(str(raw))
            except ValueError:
                warnings.append(f"Invalid {field.replace('_', ' ')} value was ignored")
    interaction["adaptive_follow_up"] = str(adaptive_follow_up).strip().lower() in {"1", "true", "yes", "on"}

    result = {
        "question_type": kind,
        "prompt": _clean(prompt),
        "instructions": _clean(instructions) or None,
        "part_heading": _clean(part_heading) or None,
        "passage": str(passage or "").strip() or None,
        "options": normalized_options,
        "correct_answers": answers,
        "interaction": interaction,
        "explanation": _clean(explanation) or None,
        "points": numeric_points,
        "difficulty": difficulty,
        "warnings": warnings,
    }
    if _clean(target_part):
        result["target_part"] = _clean(target_part)
    return result


def _parse_table_rows(
    rows: list[dict[str, str]],
    default_target_part: str = "",
    default_section_type: str = "",
) -> tuple[list[dict], list[str]]:
    questions: list[dict] = []
    warnings: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        prompt = row.get("prompt") or row.get("question") or row.get("question_text") or ""
        target_part = (
            row.get("part_code")
            or row.get("part")
            or row.get("part_title")
            or row.get("module_part")
            or default_target_part
            or default_section_type
            or ""
        )
        is_l1 = bool(
            target_part
            and (
                "listening_1" in target_part.lower()
                or "listening 1" in target_part.lower()
                or (target_part.lower().endswith("_1") and "listen" in target_part.lower())
            )
        )
        part_norm = target_part.lower().replace(" ", "_")
        is_r1b = "reading_1b" in part_norm
        is_r2 = "reading_2" in part_norm
        if not prompt and is_l1:
            prompt = f"Question {len(questions) + 1}"
        elif not prompt and is_r1b:
            r1b_count = sum(1 for q in questions if "1b" in str(q.get("target_part", "")).lower())
            prompt = f"Choose the best option for gap ({r1b_count + 1})."
        elif not prompt and is_r2:
            r2_count = sum(1 for q in questions if "reading_2" in str(q.get("target_part", "")).lower().replace(" ", "_"))
            prompt = f"Choose the best option for gap {r2_count + 1}."
        elif not prompt and not any(row.values()):
            continue
        options = []
        for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            value = row.get(f"option_{letter.lower()}") or row.get(f"choice_{letter.lower()}")
            if value:
                options.append({"key": letter, "text": value})
        if not options and row.get("options"):
            for index, value in enumerate(re.split(r"\s*\|\s*", row["options"])):
                value = OPTION_RE.sub(r"\2", value).strip()
                if value:
                    options.append({"key": chr(65 + index), "text": value})
        answer = row.get("correct_answer") or row.get("correct_answers") or row.get("answer") or ""
        preview = _question_preview(
            prompt=prompt,
            options=options,
            correct_answers=_split_answers(answer),
            question_type=row.get("question_type") or row.get("type") or "",
            instructions=row.get("instructions") or "",
            passage=row.get("passage") or row.get("context") or "",
            explanation=row.get("explanation") or row.get("rationale") or "",
            points=row.get("points") or 1,
            difficulty=row.get("difficulty") or "medium",
            group_label=row.get("group_label") or row.get("conversation") or "",
            turn_type=row.get("turn_type") or row.get("speaking_turn") or "",
            preparation_seconds=row.get("preparation_seconds"),
            response_seconds=row.get("response_seconds"),
            adaptive_follow_up=row.get("adaptive_follow_up") or False,
            target_part=target_part,
        )
        if preview["warnings"]:
            warnings.append(f"Row {row_number}: {'; '.join(preview['warnings'])}")
        questions.append(preview)
        if len(questions) >= MAX_IMPORT_QUESTIONS:
            warnings.append(f"Only the first {MAX_IMPORT_QUESTIONS} questions were extracted")
            break
    return questions, warnings


def parse_csv(
    content: bytes,
    default_target_part: str = "",
    default_section_type: str = "",
) -> tuple[str, list[dict], list[str]]:
    if isinstance(content, str):
        decoded = content
    else:
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV must be UTF-8 encoded",
            ) from exc

    try:
        dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
    if not reader.fieldnames:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV header row is missing")

    def normalized_row(raw_row: dict) -> dict[str, str]:
        result: dict[str, str] = {}
        for key, value in raw_row.items():
            norm_key = re.sub(r"[^a-z0-9]+", "_", str(key or "").strip().lower()).strip("_")
            result[norm_key] = _clean_multiline(value) if norm_key in {"passage", "context"} else _clean(value)
        return result

    rows = [normalized_row(r) for r in reader]
    questions, warnings = _parse_table_rows(
        rows,
        default_target_part=default_target_part,
        default_section_type=default_section_type,
    )
    if not questions:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No question rows were found in the CSV")
    return decoded[:MAX_EXTRACTED_TEXT], questions, warnings


def parse_excel(
    content: bytes,
    default_target_part: str = "",
    default_section_type: str = "",
) -> tuple[str, list[dict], list[str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File is not a valid Excel (.xlsx / .xls) spreadsheet",
        ) from exc

    all_rows: list[dict[str, str]] = []
    text_lines: list[str] = []

    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        header_row: Optional[list[str]] = None
        for row in rows_iter:
            if any(row):
                header_row = [
                    re.sub(r"[^a-z0-9]+", "_", str(cell or "").strip().lower()).strip("_")
                    for cell in row
                ]
                break
        if not header_row:
            continue

        recognized = {"prompt", "question", "question_text", "part_code", "part", "type", "question_type", "option_a"}
        if not any(h in recognized for h in header_row if h):
            continue

        sheet_title = sheet.title.strip()
        sheet_part = (
            sheet_title
            if any(p in sheet_title.lower() for p in ("part", "reading", "listening", "writing", "speaking"))
            else ""
        )

        for row in rows_iter:
            if not any(row):
                continue
            row_dict: dict[str, str] = {}
            for idx, cell_val in enumerate(row):
                if idx < len(header_row) and header_row[idx]:
                    key = header_row[idx]
                    if cell_val is None:
                        val_str = ""
                    elif isinstance(cell_val, float) and cell_val.is_integer():
                        val_str = str(int(cell_val))
                    else:
                        val_str = str(cell_val).strip()
                    row_dict[key] = _clean_multiline(val_str) if key in {"passage", "context"} else _clean(val_str)

            if not any(row_dict.values()):
                continue

            if not row_dict.get("part_code") and not row_dict.get("part") and sheet_part:
                row_dict["part_code"] = sheet_part

            all_rows.append(row_dict)
            prompt_text = row_dict.get("prompt") or row_dict.get("question") or ""
            if prompt_text:
                text_lines.append(f"[{sheet_title}] {prompt_text}")

    if not all_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No question rows were found in the Excel workbook. Ensure the sheet has column headers such as 'prompt', 'option_a', and 'correct_answer'.",
        )

    questions, warnings = _parse_table_rows(
        all_rows,
        default_target_part=default_target_part,
        default_section_type=default_section_type,
    )
    if not questions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid questions could be extracted from the Excel spreadsheet",
        )
    source_text = "\n".join(text_lines)
    return source_text[:MAX_EXTRACTED_TEXT], questions, warnings



def parse_pdf(
    content: bytes,
    default_target_part: str = "",
    default_section_type: str = "",
) -> tuple[str, list[dict], list[str]]:
    try:
        reader = PdfReader(io.BytesIO(content))
        pages = [(page.extract_text() or "").strip() for page in reader.pages]
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="PDF text could not be extracted") from exc
    text = "\n\n".join(page for page in pages if page)
    if not text:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No selectable text was found. Scanned PDFs need OCR before import.",
        )

    lines = [_clean(raw_line) for raw_line in text.splitlines() if _clean(raw_line)]
    answer_map: dict[int, list[str]] = {}
    content_lines = lines
    for index, line in enumerate(lines):
        if ANSWER_KEY_HEADER_RE.match(line):
            content_lines = lines[:index]
            for answer_line in lines[index + 1:]:
                match = ANSWER_KEY_ENTRY_RE.match(answer_line)
                if match:
                    answer_map[int(match.group(1))] = _split_answers(match.group(2))
            break

    questions: list[dict] = []
    warnings: list[str] = []
    current: Optional[dict] = None
    current_part = default_target_part or default_section_type or ""
    current_passage = []
    mode = "prompt"

    current_heading_lines: list[str] = []
    collecting_passage = False

    def finish() -> None:
        nonlocal current
        if not current:
            return
        norm_cp = (current.get("target_part") or current_part).lower().replace("-", "_").replace(" ", "_")
        part_head = current.get("part_heading") or (" ".join(current_heading_lines).strip()) or DEFAULT_PART_HEADINGS.get(norm_cp, "")
        is_passage_part = any(p in norm_cp for p in ("reading_1b", "reading_2", "reading_3", "reading_4", "listening_3"))
        if is_passage_part and current_passage:
            def _clean_passage_text(raw_lines: list[str]) -> str:
                if not raw_lines:
                    return ""
                raw = "\n".join(raw_lines).strip()
                lines = [l.strip() for l in raw.splitlines() if l.strip()]
                if (
                    len(lines) > 1
                    and len(lines[0]) <= 75
                    and not lines[0].endswith((".", ":", ";", "!", "?", ","))
                    and not lines[0].startswith("#")
                ):
                    raw = lines[0] + "\n\n" + "\n".join(lines[1:])
                chunks = re.split(r"\n\s*\n+", raw)
                paragraphs = []
                for chunk in chunks:
                    chunk_lines = [l.strip() for l in chunk.splitlines() if l.strip()]
                    if not chunk_lines:
                        continue
                    if re.match(r"^([#>\-*]|\d+\.)", chunk_lines[0]):
                        paragraphs.append(chunk.strip())
                    else:
                        paragraphs.append(" ".join(chunk_lines))
                return "\n\n".join(paragraphs).strip()

            passage_text = _clean_passage_text(current_passage)
            if "reading_1b" in norm_cp or "reading_2" in norm_cp:
                passage_text = _normalize_cloze_passage(passage_text)
        else:
            passage_text = None
        preview = _question_preview(
            prompt=" ".join(current["prompt"]),
            options=current["options"],
            correct_answers=current["answers"] or answer_map.get(current["number"], []),
            explanation=" ".join(current["explanation"]),
            passage=passage_text or None,
            target_part=current.get("target_part", "") or default_target_part or default_section_type or "",
        )
        if preview["warnings"]:
            warnings.append(
                f"Question {len(questions) + 1}: {'; '.join(preview['warnings'])}"
            )
        questions.append(preview)
        current = None

    for line in content_lines:
        question_match = QUESTION_RE.match(line)
        option_match = OPTION_RE.match(line)
        answer_match = ANSWER_RE.match(line)
        explanation_match = EXPLANATION_RE.match(line)
        part_match = PART_HEADER_RE.match(line)
        if part_match:
            finish()
            current_part = part_match.group(1)
            norm_cp = current_part.lower().replace("-", "_").replace(" ", "_")
            current_heading_lines = []
            current_passage = []
            collecting_passage = False
            mode = "prompt"
        elif question_match:
            finish()
            collecting_passage = False
            norm_cp = current_part.lower().replace("-", "_").replace(" ", "_")
            head_text = " ".join(current_heading_lines).strip()
            current = {
                "number": int(question_match.group(1)),
                "prompt": [question_match.group(2)],
                "options": [],
                "answers": [],
                "explanation": [],
                "part_heading": head_text or DEFAULT_PART_HEADINGS.get(norm_cp, ""),
                "target_part": current_part or default_target_part or default_section_type or "",
            }
            mode = "prompt"
        elif not current and current_part:
            clean_line = line.strip()
            if re.match(r"^(?:reading\s+)?passage(?:\s+\d+)?\s*:?$", clean_line, re.IGNORECASE):
                collecting_passage = True
            elif collecting_passage:
                current_passage.append(line)
            else:
                current_heading_lines.append(line)
        elif current and answer_match:
            current["answers"] = _split_answers(answer_match.group(1))
            mode = "answer"
        elif current and explanation_match:
            current["explanation"].append(explanation_match.group(1))
            mode = "explanation"
        elif current and option_match:
            current["options"].append(
                {"key": option_match.group(1).upper(), "text": option_match.group(2)}
            )
            mode = "option"
        elif current and mode == "option" and current["options"]:
            current["options"][-1]["text"] += f" {line}"
        elif current and mode == "explanation":
            current["explanation"].append(line)
        elif current and mode != "answer":
            current["prompt"].append(line)
    finish()

    if not questions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No numbered questions were detected. Use formats such as '1. Question' and 'A. Choice'.",
        )
    if len(questions) > MAX_IMPORT_QUESTIONS:
        questions = questions[:MAX_IMPORT_QUESTIONS]
        warnings.append(f"Only the first {MAX_IMPORT_QUESTIONS} questions were extracted")
    return text[:MAX_EXTRACTED_TEXT], questions, warnings


async def preview_upload(
    upload: UploadFile,
    default_target_part: str = "",
    default_section_type: str = "",
) -> dict:
    filename = (upload.filename or "questions").strip()
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in {"pdf", "csv", "xlsx", "xls"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Question imports must be PDF, CSV, or Excel (.xlsx / .xls) files",
        )
    content = await upload.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")
    limit = (
        MAX_PDF_BYTES
        if extension == "pdf"
        else (MAX_EXCEL_BYTES if extension in {"xlsx", "xls"} else MAX_CSV_BYTES)
    )
    if len(content) > limit:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{extension.upper()} must be {limit // 1024 // 1024} MB or smaller")
    if extension == "pdf" and not content.startswith(b"%PDF-"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is not a valid PDF")
    if extension == "csv" and b"\x00" in content[:4096]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is not a valid text CSV")

    if extension in {"xlsx", "xls"}:
        source_text, questions, warnings = parse_excel(
            content,
            default_target_part=default_target_part,
            default_section_type=default_section_type,
        )
    elif extension == "pdf":
        source_text, questions, warnings = parse_pdf(
            content,
            default_target_part=default_target_part,
            default_section_type=default_section_type,
        )
    else:
        source_text, questions, warnings = parse_csv(
            content,
            default_target_part=default_target_part,
            default_section_type=default_section_type,
        )
    return {
        "source_type": "xlsx" if extension in {"xlsx", "xls"} else extension,
        "source_filename": filename[:255],
        "source_text": source_text,
        "questions": questions,
        "question_count": len(questions),
        "warning_count": len(warnings),
        "warnings": warnings,
    }


def generate_excel_template(module_type: str = "reading") -> bytes:
    from pathlib import Path

    normalized_type = (module_type or "reading").strip().lower()
    slug = normalized_type.replace("_", "-")

    candidates = [
        Path(__file__).resolve().parents[3] / "docs" / "module-upload-templates" / "excel" / f"{slug}-full-module-upload.xlsx",
        Path(__file__).resolve().parents[3] / "docs" / "module-upload-test-files" / "excel" / f"{slug}-full-module-upload.xlsx",
        Path(__file__).resolve().parents[3] / "docs" / "module-upload-templates" / "excel" / f"{slug}-template.xlsx",
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate.read_bytes()

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Module Questions"

    columns = [
        "part_code",
        "prompt",
        "option_a",
        "option_b",
        "option_c",
        "option_d",
        "correct_answer",
        "passage",
        "instructions",
    ]

    sample_rows_by_type: dict[str, list[dict[str, object]]] = {
        "reading": [
            {
                "part_code": "Reading 1A",
                "prompt": "The committee's findings were largely **compatible** with those of the earlier study, which strengthened confidence in both sets of results.",
                "question_type": "mcq_single",
                "option_a": "consistent",
                "option_b": "persistent",
                "option_c": "insistent",
                "option_d": "resistant",
                "correct_answer": "A",
                "points": 1,
                "difficulty": "medium",
                "passage": "",
                "instructions": "Read each sentence. Choose the word that can best replace the bold word without changing the meaning.",
            },
            {
                "part_code": "Reading 1B",
                "prompt": "Choose the best option for gap (1).",
                "question_type": "mcq_single",
                "option_a": "vital",
                "option_b": "random",
                "option_c": "minor",
                "option_d": "fragile",
                "correct_answer": "A",
                "points": 1,
                "difficulty": "medium",
                "passage": "Urban green spaces play a {{blank:1}} role in cooling city streets during summer heat waves. Trees and parks reduce local temperatures while offering residents areas for recreation and social interaction.",
                "instructions": "Read the text and choose the correct word for each gap.",
            },
            {
                "part_code": "Reading 2",
                "prompt": "Choose the best option for gap 1.",
                "question_type": "matching_reusable",
                "option_a": "Researchers documented dozens of rare species during the first survey.",
                "option_b": "These ancient woodlands have survived virtually untouched for centuries.",
                "option_c": "Heavy industrial development permanently altered the surrounding terrain.",
                "correct_answer": "B",
                "points": 1,
                "difficulty": "medium",
                "passage": "The temperate rainforests of the Pacific Northwest contain some of the oldest living trees on Earth. [1] In recent decades, conservationists have worked to establish protected wildlife corridors.",
                "instructions": "Read the text. Choose the sentence that best fits each gap. One sentence is a distractor.",
            },
            {
                "part_code": "Reading 3",
                "prompt": "Which text mentions archaeological discoveries that challenged earlier historical assumptions?",
                "question_type": "matching_reusable",
                "option_a": "Text A",
                "option_b": "Text B",
                "option_c": "Text C",
                "option_d": "Text D",
                "correct_answer": "B",
                "points": 1,
                "difficulty": "medium",
                "passage": "Text A\nNew survey techniques have revealed extensive medieval settlement networks along the river valley.\n\nText B\nRecent excavations of coastal trading posts uncovered imported pottery, disproving the theory that regional communities were isolated.\n\nText C\nMuseum curators are preserving iron tools using advanced chemical treatments.\n\nText D\nUniversity researchers are digitising historical excavation field journals.",
                "instructions": "Read texts A–D. Decide which text answers each question.",
            },
            {
                "part_code": "Reading 4",
                "prompt": "What is the primary argument presented in paragraph 2 regarding renewable energy transition?",
                "question_type": "mcq_single",
                "option_a": "Storage technology must advance alongside generation capacity",
                "option_b": "Fossil fuel subsidies should be phased out immediately",
                "option_c": "Consumer demand alone is sufficient to drive transition",
                "option_d": "Nuclear power is the only reliable baseload option",
                "correct_answer": "A",
                "points": 1,
                "difficulty": "medium",
                "passage": "Transitioning national power grids to renewable sources requires solving complex intermittency challenges. While solar panels and wind turbines have become cost-competitive, utility operators must deploy grid-scale battery storage to balance fluctuating supply with peak evening demand.",
                "instructions": "Read the text and choose the correct answer for each question.",
            },
        ],
        "listening": [
            {
                "part_code": "Listening 1",
                "prompt": "Where will the orientation session take place?",
                "question_type": "mcq_single",
                "option_a": "Lecture Hall A",
                "option_b": "The main library",
                "option_c": "Student Services",
                "option_d": "Room 204",
                "correct_answer": "D",
                "points": 1,
                "difficulty": "medium",
                "instructions": "You will hear some short conversations. You will hear each conversation twice. Choose the correct answer.",
            },
            {
                "part_code": "Listening 2",
                "prompt": "Why does the student visit the academic advisor?",
                "question_type": "mcq_single",
                "option_a": "To request an extension for coursework",
                "option_b": "To discuss module prerequisites",
                "option_c": "To change degree specialisation",
                "option_d": "To apply for study abroad",
                "correct_answer": "B",
                "points": 1,
                "difficulty": "medium",
                "group_label": "Academic advising meeting",
                "instructions": "You will hear five conversations. Listen to the conversations and answer the questions. Choose the correct answer.",
            },
            {
                "part_code": "Listening 3",
                "prompt": "Archaeology research project at an Ancient Roman villa — Excavations began in the year [1].",
                "question_type": "short_answer",
                "correct_answer": "1994",
                "points": 1,
                "difficulty": "medium",
                "passage": "Archaeology research project at an Ancient Roman villa\nThe team began seasonal excavations in {{blank:1}} after geophysical scans revealed buried foundations.",
                "instructions": "You will hear a recording. You will hear the recording twice. Complete the notes with NO MORE THAN THREE WORDS for each gap.",
            },
            {
                "part_code": "Listening 4",
                "prompt": "What consensus did the panel reach regarding artificial intelligence in medical diagnostics?",
                "question_type": "mcq_single",
                "option_a": "AI should assist clinicians rather than replace them",
                "option_b": "Clinical trials are no longer required for mature models",
                "option_c": "Diagnostic algorithms should remain proprietary",
                "correct_answer": "A",
                "points": 1,
                "difficulty": "medium",
                "instructions": "You will hear a discussion. You will hear the discussion twice. Choose the correct answer for each question.",
            },
        ],
        "writing": [
            {
                "part_code": "Writing 1",
                "prompt": "The bar chart shows international student enrollment across four academic departments from 2020 to 2024. Summarise the information by selecting and reporting the main features, and make comparisons where relevant.",
                "question_type": "essay",
                "points": 32,
                "difficulty": "medium",
                "instructions": "Write at least 150 words.",
                "passage": "Business: 2020 (420), 2024 (580). Engineering: 2020 (380), 2024 (510). Arts: 2020 (210), 2024 (190). Sciences: 2020 (310), 2024 (440).",
            },
            {
                "part_code": "Writing 2",
                "prompt": "Some people argue that universities should focus exclusively on job-specific training, while others believe higher education should provide a broad general knowledge. Discuss both views and give your own opinion.",
                "question_type": "essay",
                "points": 32,
                "difficulty": "medium",
                "instructions": "Write at least 250 words.",
            },
        ],
        "speaking": [
            {
                "part_code": "Speaking 1",
                "prompt": "Please state your full name and candidate number.",
                "question_type": "speaking_prompt",
                "points": 1,
                "difficulty": "easy",
                "turn_type": "identity",
                "preparation_seconds": 0,
                "response_seconds": 30,
            },
            {
                "part_code": "Speaking 1",
                "prompt": "What hobbies or activities do you enjoy in your free time?",
                "question_type": "speaking_prompt",
                "points": 1,
                "difficulty": "medium",
                "turn_type": "topic_question",
                "preparation_seconds": 0,
                "response_seconds": 45,
            },
            {
                "part_code": "Speaking 2",
                "prompt": "You are arranging a campus study group. Suggest meeting times to your classmate and explain your preference.",
                "question_type": "speaking_prompt",
                "points": 1,
                "difficulty": "medium",
                "turn_type": "situation_interaction",
                "preparation_seconds": 20,
                "response_seconds": 90,
            },
            {
                "part_code": "Speaking 3",
                "prompt": "Give a short presentation on a book or article that made a strong impression on you. Explain why it was memorable.",
                "question_type": "speaking_prompt",
                "points": 1,
                "difficulty": "medium",
                "turn_type": "topic_presentation",
                "preparation_seconds": 60,
                "response_seconds": 120,
            },
            {
                "part_code": "Speaking 4",
                "prompt": "How has digital communication influenced interpersonal relationships in modern society?",
                "question_type": "speaking_prompt",
                "points": 1,
                "difficulty": "hard",
                "turn_type": "academic_discussion",
                "preparation_seconds": 30,
                "response_seconds": 120,
            },
        ],
    }

    norm_type = (module_type or "reading").lower().replace("-", "_")
    if norm_type in {"final_test", "full_mock"}:
        rows = (
            sample_rows_by_type["reading"]
            + sample_rows_by_type["listening"]
            + sample_rows_by_type["writing"]
            + sample_rows_by_type["speaking"]
        )
    else:
        rows = sample_rows_by_type.get(norm_type, sample_rows_by_type["reading"])

    header_fill = PatternFill(start_color="B80F28", end_color="B80F28", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )

    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row_dict in enumerate(rows, start=2):
        row_vals = [str(row_dict.get(c, "") or "") for c in columns]
        ws.append(row_vals)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = {
        "A": 16,  # part_code
        "B": 48,  # prompt
        "C": 18,  # question_type
        "D": 22,  # option_a
        "E": 22,  # option_b
        "F": 22,  # option_c
        "G": 22,  # option_d
        "H": 16,  # correct_answer
        "I": 10,  # points
        "J": 12,  # difficulty
        "K": 45,  # passage
        "L": 35,  # instructions
        "M": 25,  # group_label
        "N": 20,  # turn_type
        "O": 18,  # prep seconds
        "P": 18,  # resp seconds
        "Q": 30,  # explanation
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 26
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

