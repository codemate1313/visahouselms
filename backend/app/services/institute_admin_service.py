import csv
import io
import secrets
import string
from zipfile import BadZipFile
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import EmailStr, TypeAdapter, ValidationError
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload, selectinload

from app.core.security import hash_password
from app.dependencies.limits import enforce_limit
from app.models.attempt import AttemptPartGrade, TestAttempt
from app.models.audit_log import AuditLog
from app.models.exam_module import ExamModule, InstituteModule
from app.models.institute import Institute
from app.models.role import DEVELOPER, INST_INSTRUCTOR, INSTITUTE_ADMIN, STUDENT, SUPER_ADMIN, Role
from app.models.user import (
    ACCESS_ACTIVE,
    ACCESS_EXPIRED,
    ACCESS_RELEASED,
    ACCESS_SUSPENDED,
    SEAT_HOLDING_STATES,
    User,
    seat_holder_filter,
)
from app.models.user_device import UserDevice
from app.models.user_session import UserSession
from app.services import (
    access_window_service,
    account_service,
    institute_service,
    notification_service,
    subscription_service,
)

MANAGED_ROLES = (INST_INSTRUCTOR, STUDENT)
MAX_IMPORT_ROWS = 1000
EMAIL_ADAPTER = TypeAdapter(EmailStr)


def _temporary_password() -> str:
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    chars = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%"),
        *(secrets.choice(alphabet) for _ in range(10)),
    ]
    secrets.SystemRandom().shuffle(chars)
    return "".join(chars)


def _require_institute(actor: User, scoped_institute_id: Optional[int] = None) -> int:
    if actor.role.name in (SUPER_ADMIN, DEVELOPER):
        if scoped_institute_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="An institute is required for this operation",
            )
        return scoped_institute_id
    if actor.institute_id is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account is not associated with an institute",
        )
    if scoped_institute_id is not None and scoped_institute_id != actor.institute_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")
    return actor.institute_id


def admin_permissions(actor: User) -> dict:
    if actor.institute is None:
        return institute_service.normalized_admin_permissions(None)
    return institute_service.normalized_admin_permissions(actor.institute.admin_permissions)


def require_admin_permission(actor: User, *permissions: str) -> None:
    if actor.role.name == SUPER_ADMIN:
        return
    granted = admin_permissions(actor)
    if not any(granted.get(permission, False) for permission in permissions):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="This activity has not been permitted by the Super Admin",
        )


def _audit(
    db: Session,
    actor: User,
    action: str,
    member_id: Optional[int],
    ip: Optional[str],
    details: Optional[dict] = None,
) -> None:
    db.add(
        AuditLog(
            user_id=actor.id,
            action=action,
            entity_type="institute_member",
            entity_id=member_id,
            details=details,
            ip_address=ip,
        )
    )


def _member_query(db: Session, institute_id: int, *, include_deleted: bool = False):
    """Backs the roster and every by-id lookup.

    Retired members are excluded by default, so a deleted account is neither
    listed nor reachable through a link a browser tab still remembers. The
    exception is the roster's own "Deleted" filter, which exists precisely to
    show them - hence the opt-out rather than a filter hard-wired in here.
    """
    query = (
        db.query(User)
        .options(joinedload(User.role))
        .join(Role, User.role_id == Role.id)
        .filter(User.institute_id == institute_id, Role.name.in_(MANAGED_ROLES))
    )
    if not include_deleted:
        query = query.filter(User.deleted_at.is_(None))
    return query


def serialize_member(user: User, metrics: Optional[dict] = None) -> dict:
    metrics = metrics or {}
    return {
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "role": user.role.name,
        "is_active": user.is_active,
        "force_password_reset": user.force_password_reset,
        "phone_number": user.phone_number,
        "address": user.address,
        "avatar_url": account_service.avatar_url_for(user),
        "deleted_at": user.deleted_at,
        "attempt_count": metrics.get("attempt_count", 0),
        "device_count": metrics.get("device_count", 0),
        "active_session_count": metrics.get("active_session_count", 0),
        "created_at": user.created_at,
        # The window and the seat. `is_active` above still answers "can they log
        # in"; these answer "until when" and "does this cost a seat", which the
        # roster needs to show and the old boolean could not express.
        **access_window_service.serialize_access(user),
    }


def _metrics_for_members(db: Session, member_ids: list[int]) -> dict[int, dict]:
    metrics = {
        member_id: {"attempt_count": 0, "device_count": 0, "active_session_count": 0}
        for member_id in member_ids
    }
    if not member_ids:
        return metrics

    for user_id, count in (
        db.query(TestAttempt.user_id, func.count(TestAttempt.id))
        .filter(TestAttempt.user_id.in_(member_ids))
        .group_by(TestAttempt.user_id)
        .all()
    ):
        metrics[user_id]["attempt_count"] = count
    for user_id, count in (
        db.query(UserDevice.user_id, func.count(UserDevice.id))
        .filter(UserDevice.user_id.in_(member_ids))
        .group_by(UserDevice.user_id)
        .all()
    ):
        metrics[user_id]["device_count"] = count
    now = datetime.now(timezone.utc)
    for user_id, count in (
        db.query(UserSession.user_id, func.count(UserSession.id))
        .filter(
            UserSession.user_id.in_(member_ids),
            UserSession.revoked_at.is_(None),
            UserSession.expires_at > now,
        )
        .group_by(UserSession.user_id)
        .all()
    ):
        metrics[user_id]["active_session_count"] = count
    return metrics


def list_members(
    db: Session,
    actor: User,
    role_name: Optional[str] = None,
    search: Optional[str] = None,
    active: Optional[bool] = None,
    status_filter: Optional[str] = None,
    has_attempts: Optional[bool] = None,
    has_devices: Optional[bool] = None,
    has_active_sessions: Optional[bool] = None,
    scoped_institute_id: Optional[int] = None,
) -> list[dict]:
    institute_id = _require_institute(actor, scoped_institute_id)
    query = _member_query(db, institute_id, include_deleted=status_filter == "deleted")
    if role_name is not None:
        if role_name not in MANAGED_ROLES:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid member role")
        query = query.filter(Role.name == role_name)
    if search and search.strip():
        term = f"%{search.strip()}%"
        query = query.filter(
            or_(User.email.ilike(term), User.first_name.ilike(term), User.last_name.ilike(term))
        )
    if active is not None:
        query = query.filter(User.is_active.is_(active))
    if status_filter:
        if status_filter == "active":
            query = query.filter(User.is_active.is_(True), User.deleted_at.is_(None))
        elif status_filter == "inactive":
            query = query.filter(User.is_active.is_(False), User.deleted_at.is_(None))
        elif status_filter == "deleted":
            query = query.filter(User.deleted_at.is_not(None))
        elif status_filter == "password_reset":
            query = query.filter(User.force_password_reset.is_(True), User.deleted_at.is_(None))
        elif status_filter == "expired":
            query = query.filter(User.access_state == ACCESS_EXPIRED, User.deleted_at.is_(None))
        elif status_filter == "released":
            # "Past students" - the returning-student list. Their records and
            # emails are intact; they simply hold no seat.
            query = query.filter(User.access_state == ACCESS_RELEASED, User.deleted_at.is_(None))
        elif status_filter == "reclaimable":
            # Everyone whose seat could be freed right now: the admin's shortlist
            # when they are at their cap.
            query = query.filter(
                User.access_state.in_((ACCESS_EXPIRED, ACCESS_SUSPENDED)),
                User.deleted_at.is_(None),
            )
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid member status filter")
    users = query.order_by(User.created_at.desc()).all()
    metrics = _metrics_for_members(db, [user.id for user in users])
    rows = [serialize_member(user, metrics[user.id]) for user in users]
    if has_attempts is not None:
        rows = [row for row in rows if (row["attempt_count"] > 0) is has_attempts]
    if has_devices is not None:
        rows = [row for row in rows if (row["device_count"] > 0) is has_devices]
    if has_active_sessions is not None:
        rows = [row for row in rows if (row["active_session_count"] > 0) is has_active_sessions]
    return rows


def member_capacity(db: Session, actor: User, scoped_institute_id: Optional[int] = None) -> dict:
    institute_id = _require_institute(actor, scoped_institute_id)
    institute = db.get(Institute, institute_id)
    if institute is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Institute not found")

    counts = subscription_service.usage(db, institute_id)
    limits = {
        "students": institute.student_limit,
        "staff": institute.staff_limit,
    }
    subscription, state = subscription_service.current_subscription(db, institute_id)
    if institute.onboarding_status != "draft" and subscription is not None and subscription.plan is not None:
        # Summed across every live term, not read off the one "current" row, so
        # an institute running two plans sees the capacity of both. Same helper
        # `enforce_limit` uses, so the number on this screen and the number that
        # blocks Add student are the same number.
        from app.dependencies.limits import plan_limit_total

        limits = {
            "students": plan_limit_total(db, institute_id, "student_limit"),
            "staff": plan_limit_total(db, institute_id, "staff_limit"),
        }

    can_add = {
        resource: state in ("active", "grace") and limit is not None and counts[resource] < limit
        for resource, limit in limits.items()
    }
    if institute.onboarding_status == "draft":
        can_add = {
            resource: limit is not None and counts[resource] < limit
            for resource, limit in limits.items()
        }

    # A breakdown of who is holding the seats, so the admin can see at a glance
    # how many are reclaimable without reading the whole roster.
    seat_states = dict(
        db.query(User.access_state, func.count(User.id))
        .join(Role, User.role_id == Role.id)
        .filter(
            User.institute_id == institute_id,
            Role.name == STUDENT,
            seat_holder_filter(),
        )
        .group_by(User.access_state)
        .all()
    )
    released_students = (
        db.query(User)
        .join(Role, User.role_id == Role.id)
        .filter(
            User.institute_id == institute_id,
            Role.name == STUDENT,
            User.access_state == ACCESS_RELEASED,
            User.deleted_at.is_(None),
        )
        .count()
    )

    student_limit = limits["students"]
    return {
        "usage": {
            "students": counts["students"],
            "staff": counts["staff"],
        },
        "limits": limits,
        "can_add": can_add,
        "seats": {
            "used": counts["students"],
            "total": student_limit,
            "free": max(0, student_limit - counts["students"]) if student_limit is not None else None,
            "active": seat_states.get(ACCESS_ACTIVE, 0),
            "suspended": seat_states.get(ACCESS_SUSPENDED, 0),
            # Holding a seat but unable to log in - the ones worth reclaiming.
            "expired": seat_states.get(ACCESS_EXPIRED, 0),
            "reclaimable": seat_states.get(ACCESS_EXPIRED, 0) + seat_states.get(ACCESS_SUSPENDED, 0),
            "past_students": released_students,
        },
        "subscription_ends_on": (
            access_window_service.to_local_date(
                access_window_service.subscription_ceiling(db, institute_id),
                access_window_service.institute_timezone(institute),
            ).isoformat()
            if access_window_service.subscription_ceiling(db, institute_id) is not None
            else None
        ),
    }


def get_member_or_404(
    db: Session,
    actor: User,
    member_id: int,
    scoped_institute_id: Optional[int] = None,
) -> User:
    user = _member_query(db, _require_institute(actor, scoped_institute_id)).filter(User.id == member_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")
    return user


def _lock_institute_seats(db: Session, institute_id: int) -> None:
    """Hold the seat count still between counting it and inserting a row.

    `enforce_limit` counts, returns, and only then does the caller insert. With
    nothing held in between, two admins clicking Add student in the same second
    - or one double-clicking - both read 99 of 100, both pass the check, and
    both insert. 101 seats on a 100-seat plan, and no error anywhere.

    Locking the institute row serialises the whole count-then-insert, so the
    second request waits, re-counts 100, and is refused. On SQLite the
    connection-level write lock already gives this for free and FOR UPDATE is
    not supported, so it is skipped there.
    """
    if db.bind is None or db.bind.dialect.name == "sqlite":
        return
    from app.models.institute import Institute

    db.query(Institute).filter(Institute.id == institute_id).with_for_update().first()


def create_member(
    db: Session,
    actor: User,
    *,
    email: str,
    first_name: str,
    last_name: str,
    role_name: str,
    phone_number: Optional[str],
    address: Optional[str],
    ip: Optional[str],
    access_starts_on: Optional[date] = None,
    access_ends_on: Optional[date] = None,
    scoped_institute_id: Optional[int] = None,
) -> dict:
    institute_id = _require_institute(actor, scoped_institute_id)
    if role_name not in MANAGED_ROLES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid member role")

    normalized_email = account_service.ensure_user_credentials_available(db, email)

    # Students carry an access window and it is never defaulted. A default is
    # how a student ends up outliving the subscription that paid for them.
    access_starts_at = access_ends_at = None
    if role_name == STUDENT:
        if access_starts_on is None or access_ends_on is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A student needs an access start and end date.",
            )
        access_starts_at, access_ends_at = access_window_service.resolve_window(
            db, institute_id, access_starts_on, access_ends_on
        )

    _lock_institute_seats(db, institute_id)
    enforce_limit(db, institute_id, "students" if role_name == STUDENT else "staff")
    role = db.query(Role).filter(Role.name == role_name).first()
    if role is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"{role_name} role is not seeded",
        )

    temporary_password = _temporary_password()
    from app.models.institute import Institute
    institute = db.get(Institute, institute_id)
    user = User(
        email=normalized_email,
        password_hash=hash_password(temporary_password),
        role_id=role.id,
        institute_id=institute_id,
        first_name=first_name,
        last_name=last_name,
        phone_number=phone_number,
        address=address,
        is_active=institute is None or institute.onboarding_status != "draft",
        force_password_reset=True,
        access_starts_at=access_starts_at,
        access_ends_at=access_ends_at,
        access_state=ACCESS_ACTIVE,
    )
    db.add(user)
    db.flush()
    _audit(
        db,
        actor,
        "institute_member.create",
        user.id,
        ip,
        {
            "email": user.email,
            "role": role_name,
            "access_starts_on": access_starts_on.isoformat() if access_starts_on else None,
            "access_ends_on": access_ends_on.isoformat() if access_ends_on else None,
        },
    )
    db.commit()
    account_service.send_account_credentials_email(
        db, user, temporary_password, role_label=role_name.replace("_", " ").title()
    )
    result = serialize_member(get_member_or_404(db, actor, user.id, scoped_institute_id))
    result["temporary_password"] = temporary_password
    return result


def update_member(
    db: Session,
    actor: User,
    member_id: int,
    data: dict,
    fields_set: set[str],
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> dict:
    user = get_member_or_404(db, actor, member_id, scoped_institute_id)
    if user.deleted_at is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Archived members cannot be edited")
    if "email" in fields_set and data.get("email") is not None:
        candidate_email = str(data["email"])
        if account_service.normalize_email(candidate_email) != account_service.normalize_email(user.email):
            user.email = account_service.ensure_user_credentials_available(
                db, candidate_email, exclude_user_id=user.id
            )
    for field in ("first_name", "last_name", "phone_number", "address"):
        if field in fields_set:
            setattr(user, field, data.get(field))

    db.add(user)
    _audit(db, actor, "institute_member.update", user.id, ip, {"fields": sorted(fields_set)})
    db.commit()
    return serialize_member(get_member_or_404(db, actor, user.id, scoped_institute_id))


def set_member_active(
    db: Session,
    actor: User,
    member_id: int,
    active: bool,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> dict:
    user = _member_query(db, _require_institute(actor, scoped_institute_id), include_deleted=True).filter(User.id == member_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")
    if user.deleted_at is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Archived members cannot be reactivated")

    # Deactivating never releases a seat - that is a separate, deliberate action
    # (`release_seat`). Turning someone back on therefore costs nothing and
    # needs no limit check: they never stopped holding their seat.
    if user.access_state == ACCESS_RELEASED:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This student's seat has been released. Use Reactivate to give "
                "them a new access window and a seat."
            ),
        )

    if active and user.role.name == STUDENT and not access_window_service.is_window_open(user):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This student's access window has ended. Extend the window to "
                "let them back in."
            ),
        )

    user.is_active = active
    if user.role.name == STUDENT:
        user.access_state = ACCESS_ACTIVE if active else ACCESS_SUSPENDED
    revoked = account_service.revoke_all_sessions(db, user.id) if not active else 0
    db.add(user)
    _audit(
        db,
        actor,
        "institute_member.reactivate" if active else "institute_member.deactivate",
        user.id,
        ip,
        {"sessions_revoked": revoked},
    )
    db.commit()
    notification_service.send_account_status_email(db, user, active)
    return serialize_member(get_member_or_404(db, actor, user.id, scoped_institute_id))


def release_seat(
    db: Session,
    actor: User,
    member_id: int,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> dict:
    """Hand a student's seat back to the institute, keeping everything else.

    This is the only action that reduces the seat count without destroying the
    student. Their row, email, attempts, results and history all stay exactly
    where they are, and they remain searchable, so a returner can be found and
    reactivated. Deleting is still available and still does what it always did;
    this exists so an admin never has to choose between reclaiming a seat and
    keeping the record of the student who had it.

    Refused while the student can still log in. Freeing a seat from under
    someone mid-course should take two deliberate steps, not one misclick on a
    roster row - so they must be expired or deactivated first.
    """
    institute_id = _require_institute(actor, scoped_institute_id)
    user = _member_query(db, institute_id).filter(User.id == member_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")
    if user.role.name != STUDENT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only student seats can be released.",
        )
    if user.access_state == ACCESS_RELEASED:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="This student's seat has already been released.",
        )
    if user.access_state == ACCESS_ACTIVE:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This student still has access. Deactivate them, or wait for "
                "their access window to end, before releasing the seat."
            ),
        )

    # A student mid-exam keeps their seat until the sitting is over; releasing
    # would log them out on their next autosave and lose the recording.
    if access_window_service.students_with_live_attempts(db, [user.id]):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="This student has a test in progress. Try again once it is submitted.",
        )

    previous_state = user.access_state
    user.access_state = ACCESS_RELEASED
    user.is_active = False
    revoked = account_service.revoke_all_sessions(db, user.id)
    db.add(user)
    _audit(
        db,
        actor,
        "institute_member.seat_released",
        user.id,
        ip,
        {"previous_state": previous_state, "sessions_revoked": revoked},
    )
    db.commit()
    return serialize_member(get_member_or_404(db, actor, user.id, scoped_institute_id))


def reactivate_seat(
    db: Session,
    actor: User,
    member_id: int,
    *,
    access_starts_on: date,
    access_ends_on: date,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> dict:
    """Bring a past student back into a seat, with a new access window.

    Reactivating costs a seat, so it goes through exactly the same limit check
    as creating a student. Without that check an admin could deactivate ten
    students, create ten more, then reactivate the original ten and sit at 110
    on a 100-seat plan.

    A new window is required rather than optional. Restoring the old one would
    reactivate a student straight back into a date that has already passed - the
    admin sees a success message, the roster shows them as active, and the
    student still cannot log in.
    """
    institute_id = _require_institute(actor, scoped_institute_id)
    user = _member_query(db, institute_id).filter(User.id == member_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")
    if user.role.name != STUDENT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only students hold seats that can be reactivated.",
        )
    if user.access_state != ACCESS_RELEASED:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This student still holds a seat. Use Extend access to change "
                "their dates instead."
            ),
        )

    starts_at, ends_at = access_window_service.resolve_window(
        db, institute_id, access_starts_on, access_ends_on
    )

    # Taking a seat back out of the pool - same gate as creating a student.
    _lock_institute_seats(db, institute_id)
    enforce_limit(db, institute_id, "students")

    user.access_starts_at = starts_at
    user.access_ends_at = ends_at
    user.access_state = ACCESS_ACTIVE
    user.is_active = True
    db.add(user)
    _audit(
        db,
        actor,
        "institute_member.seat_reactivated",
        user.id,
        ip,
        {
            "access_starts_on": access_starts_on.isoformat(),
            "access_ends_on": access_ends_on.isoformat(),
        },
    )
    db.commit()
    return serialize_member(get_member_or_404(db, actor, user.id, scoped_institute_id))


def set_member_window(
    db: Session,
    actor: User,
    member_id: int,
    *,
    access_starts_on: date,
    access_ends_on: date,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> dict:
    """Change a seated student's access dates.

    Costs nothing, because the student never gave the seat up - which is exactly
    why moving a date must never be able to free one. If an expired end date
    released a seat, this function would be a seat printer: shorten the window,
    let the sweep free the seat, fill it, then lengthen the window again.
    """
    institute_id = _require_institute(actor, scoped_institute_id)
    user = _member_query(db, institute_id).filter(User.id == member_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")
    if user.role.name != STUDENT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only students have access windows.",
        )
    if user.access_state == ACCESS_RELEASED:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="This student holds no seat. Reactivate them to grant access again.",
        )

    starts_at, ends_at = access_window_service.resolve_window(
        db, institute_id, access_starts_on, access_ends_on, allow_past_start=True
    )

    previous = {
        "access_starts_at": user.access_starts_at.isoformat() if user.access_starts_at else None,
        "access_ends_at": user.access_ends_at.isoformat() if user.access_ends_at else None,
        "access_state": user.access_state,
    }
    user.access_starts_at = starts_at
    user.access_ends_at = ends_at

    # An extension puts an expired student straight back to work; a suspension
    # is an admin decision and is left alone.
    if user.access_state == ACCESS_EXPIRED and access_window_service.is_window_open(user):
        user.access_state = ACCESS_ACTIVE
        user.is_active = True

    db.add(user)
    _audit(
        db,
        actor,
        "institute_member.window_changed",
        user.id,
        ip,
        {
            "from": previous,
            "access_starts_on": access_starts_on.isoformat(),
            "access_ends_on": access_ends_on.isoformat(),
        },
    )
    db.commit()
    return serialize_member(get_member_or_404(db, actor, user.id, scoped_institute_id))


def reset_member_password(
    db: Session,
    actor: User,
    member_id: int,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> str:
    user = get_member_or_404(db, actor, member_id, scoped_institute_id)
    if user.deleted_at is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Archived members cannot reset passwords")
    temporary_password = _temporary_password()
    user.password_hash = hash_password(temporary_password)
    user.force_password_reset = True
    user.password_changed_at = datetime.now(timezone.utc)
    revoked = account_service.revoke_all_sessions(db, user.id)
    db.add(user)
    _audit(db, actor, "institute_member.reset_password", user.id, ip, {"sessions_revoked": revoked})
    db.commit()

    try:
        from app.config import settings
        from app.services import email_template_service, smtp_service

        login_url = f"{settings.frontend_url.rstrip('/')}/login"
        first_name = user.first_name or "Member"
        subject, plain, html = email_template_service.render_password_reset_by_admin_email(
            first_name=first_name,
            email=user.email,
            new_password=temporary_password,
            login_url=login_url,
        )
        smtp_service.send_email(db, user.email, subject, plain, html)
    except Exception:
        pass
    return temporary_password


def delete_member(
    db: Session,
    actor: User,
    member_id: int,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> None:
    user = get_member_or_404(db, actor, member_id, scoped_institute_id)
    deleted_email = user.email
    deleted_role = user.role.name if user.role else None
    deleted_institute_id = user.institute_id
    _audit(
        db,
        actor,
        "institute_member.delete",
        user.id,
        ip,
        {"email": deleted_email, "role": deleted_role},
    )
    # Retires the row rather than removing it; see soft_delete_user. The audit
    # entry above is written first, so the real address is preserved before the
    # account's own copy is released.
    account_service.soft_delete_user(db, user)
    db.commit()


def revoke_member_sessions(
    db: Session,
    actor: User,
    member_id: int,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> int:
    user = get_member_or_404(db, actor, member_id, scoped_institute_id)
    revoked = account_service.revoke_all_sessions(db, user.id)
    _audit(db, actor, "institute_member.revoke_sessions", user.id, ip, {"count": revoked})
    db.commit()
    notification_service.create_notification(
        db,
        user_id=user.id,
        kind="account_sessions_revoked",
        title="Sessions revoked",
        message="Your active sessions were revoked by an administrator.",
        link_url="/notifications",
    )
    return revoked


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace("-", "_").replace(" ", "_")


def _rows_from_upload(content: bytes, filename: str) -> list[dict]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV files must use UTF-8 encoding") from exc
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The CSV file has no header row")
        return [
            {_normalize_header(key): value for key, value in row.items() if key is not None}
            for row in reader
            if any(str(value or "").strip() for value in row.values())
        ]
    if suffix == "xlsx":
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [_normalize_header(value) for value in next(values)]
            rows = [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                for row in values
                if any(str(value or "").strip() for value in row)
            ]
            workbook.close()
            return rows
        except (BadZipFile, InvalidFileException, StopIteration, ValueError, OSError) as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unable to read the Excel workbook") from exc
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload a .csv or .xlsx file")


def _value(row: dict, *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


# Accepted date spellings in an uploaded file. ISO first because it is
# unambiguous; the day-first forms are what an Indian institute's spreadsheet
# actually produces. Month-first is deliberately absent - 03/04/2027 cannot be
# both April 3rd and March 4th, and guessing wrong grants or denies a month of
# access silently.
_IMPORT_DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y", "%d %B %Y")


def _parse_import_date(value: object) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # openpyxl hands back datetimes for real date cells, strings for text cells.
    for fmt in _IMPORT_DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _import_window(row: dict) -> tuple[Optional[date], Optional[date], Optional[str]]:
    """The access dates on one import row, or a reason the row cannot be used."""
    raw_start = _value(row, "access_start", "access_starts_on", "start_date", "access_from")
    raw_end = _value(row, "access_end", "access_ends_on", "end_date", "access_to")
    if not raw_start or not raw_end:
        return None, None, "Access start and end dates are required"
    starts_on = _parse_import_date(raw_start)
    ends_on = _parse_import_date(raw_end)
    if starts_on is None:
        return None, None, f"Could not read the access start date '{raw_start}' (use YYYY-MM-DD)"
    if ends_on is None:
        return None, None, f"Could not read the access end date '{raw_end}' (use YYYY-MM-DD)"
    if ends_on < starts_on:
        return None, None, "Access end date is before the start date"
    return starts_on, ends_on, None


def _import_identity(row: dict) -> tuple[str, str, str, Optional[str], Optional[str]]:
    email = _value(row, "email", "email_address").lower()
    first_name = _value(row, "first_name", "firstname", "given_name")
    last_name = _value(row, "last_name", "lastname", "surname", "family_name")
    if not first_name or not last_name:
        full_name = _value(row, "name", "full_name", "student_name")
        parts = full_name.split()
        if parts:
            first_name = first_name or parts[0]
            last_name = last_name or (" ".join(parts[1:]) if len(parts) > 1 else "-")
    phone = _value(row, "phone_number", "phone", "mobile", "mobile_number") or None
    address = _value(row, "address") or None
    return email, first_name, last_name, phone, address


def _available_student_slots(db: Session, institute_id: int) -> int:
    from app.models.institute import Institute

    institute = db.get(Institute, institute_id)
    if institute is not None and institute.onboarding_status == "draft":
        current = (
            db.query(User)
            .join(Role, User.role_id == Role.id)
            .filter(
                User.institute_id == institute_id,
                seat_holder_filter(),
                Role.name == STUDENT,
            )
            .count()
        )
        return max(0, (institute.student_limit or 0) - current)
    subscription = subscription_service.subscription_status(db, institute_id)
    if subscription["state"] not in ("active", "grace") or not subscription["limits"]:
        raise HTTPException(
            status_code=status.HTTP_402_PAYMENT_REQUIRED,
            detail="This institute has no active subscription. Purchase or renew a plan to add students.",
        )
    # Capacity is the sum across live terms - a second plan adds seats. Usage
    # comes from the shared seat rule, so import, creation and the roster panel
    # cannot disagree.
    from app.dependencies.limits import _count_students, plan_limit_total

    return max(0, plan_limit_total(db, institute_id, "student_limit") - _count_students(db, institute_id))


def import_students(
    db: Session,
    actor: User,
    *,
    content: bytes,
    filename: str,
    ip: Optional[str],
    scoped_institute_id: Optional[int] = None,
) -> dict:
    institute_id = _require_institute(actor, scoped_institute_id)
    from app.models.institute import Institute
    institute = db.get(Institute, institute_id)
    rows = _rows_from_upload(content, filename)
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The import file contains no student rows")
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"A single import can contain at most {MAX_IMPORT_ROWS} rows")

    role = db.query(Role).filter(Role.name == STUDENT).first()
    if role is None:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="STUDENT role is not seeded")

    # One lock for the whole import, so a concurrent Add student cannot slip
    # a row in between our count and our last insert.
    _lock_institute_seats(db, institute_id)
    available = _available_student_slots(db, institute_id)
    seen: set[str] = set()
    created: list[dict] = []
    created_users: list[tuple[User, str]] = []
    skipped: list[dict] = []

    for row_number, row in enumerate(rows, start=2):
        email, first_name, last_name, phone, address = _import_identity(row)
        reason = None
        invalid_email = False
        if not email or not first_name or not last_name:
            reason = "Email and student name are required"
            invalid_email = not email
        else:
            try:
                email = str(EMAIL_ADAPTER.validate_python(email)).lower()
                email = account_service.validate_account_email(email)
            except (ValidationError, HTTPException):
                reason = account_service.INVALID_ACCOUNT_EMAIL_DETAIL
                invalid_email = True
        if reason is None and email in seen:
            reason = "Duplicate email in file"
        seen.add(email)
        if reason is None and db.query(User).filter(func.lower(User.email) == email).first() is not None:
            reason = account_service.USER_CREDENTIALS_CONFLICT_DETAIL

        starts_on = ends_on = None
        if reason is None:
            starts_on, ends_on, window_error = _import_window(row)
            reason = window_error
        if reason is None:
            try:
                access_starts_at, access_ends_at = access_window_service.resolve_window(
                    db, institute_id, starts_on, ends_on
                )
            except HTTPException as exc:
                # A window the institute cannot grant - past its subscription,
                # already expired. Skip the row and say exactly why, rather than
                # failing the whole upload on one bad date.
                reason = exc.detail
        if reason is None and len(created) >= available:
            reason = "Student plan limit reached"
        if reason is not None:
            skipped.append(
                {
                    "row": row_number,
                    "email": email or None,
                    "reason": reason,
                    "invalid_email": invalid_email,
                }
            )
            continue

        temporary_password = _temporary_password()
        user = User(
            email=email,
            password_hash=hash_password(temporary_password),
            role_id=role.id,
            institute_id=institute_id,
            first_name=first_name[:100],
            last_name=last_name[:100],
            phone_number=phone[:50] if phone else None,
            address=address[:255] if address else None,
            is_active=institute is None or institute.onboarding_status != "draft",
            force_password_reset=True,
            access_starts_at=access_starts_at,
            access_ends_at=access_ends_at,
            access_state=ACCESS_ACTIVE,
        )
        db.add(user)
        db.flush()
        _audit(
            db,
            actor,
            "institute_member.import",
            user.id,
            ip,
            {
                "row": row_number,
                "access_starts_on": starts_on.isoformat(),
                "access_ends_on": ends_on.isoformat(),
            },
        )
        created.append(
            {
                "id": user.id,
                "row": row_number,
                "email": email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "temporary_password": temporary_password,
                "access_starts_on": starts_on.isoformat(),
                "access_ends_on": ends_on.isoformat(),
            }
        )
        created_users.append((user, temporary_password))

    db.commit()
    for created_user, created_password in created_users:
        account_service.send_account_credentials_email(db, created_user, created_password, role_label="Student")
    invalid_emails = [
        {"row": item["row"], "email": item["email"], "reason": item["reason"]}
        for item in skipped
        if item.get("invalid_email")
    ]
    return {
        "summary": {
            "total_rows": len(rows),
            "created": len(created),
            "skipped": len(skipped),
            "invalid_emails": len(invalid_emails),
            "remaining_slots": max(0, available - len(created)),
        },
        "created": created,
        "skipped": skipped,
        "invalid_emails": invalid_emails,
    }


def student_overview(
    db: Session,
    actor: User,
    student_id: int,
    scoped_institute_id: Optional[int] = None,
) -> dict:
    student = get_member_or_404(db, actor, student_id, scoped_institute_id)
    if student.role.name != STUDENT:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")

    now = datetime.now(timezone.utc)
    active_sessions = (
        db.query(UserSession)
        .filter(
            UserSession.user_id == student.id,
            UserSession.revoked_at.is_(None),
            UserSession.expires_at > now,
        )
        .all()
    )
    active_device_ids = {session.device_id for session in active_sessions if session.device_id is not None}
    devices = (
        db.query(UserDevice)
        .filter(UserDevice.user_id == student.id)
        .order_by(UserDevice.last_seen_at.desc())
        .all()
    )
    attempts = (
        db.query(TestAttempt)
        .options(
            joinedload(TestAttempt.module),
            selectinload(TestAttempt.part_grades).joinedload(AttemptPartGrade.part),
            selectinload(TestAttempt.part_grades).joinedload(AttemptPartGrade.grader),
        )
        .filter(TestAttempt.user_id == student.id)
        .order_by(TestAttempt.started_at.desc())
        .all()
    )
    metrics = _metrics_for_members(db, [student.id])[student.id]
    return {
        "student": serialize_member(student, metrics),
        "security": {
            "device_count": len(devices),
            "active_session_count": len(active_sessions),
            "last_login_at": devices[0].last_seen_at if devices else None,
            "devices": [
                {
                    "id": device.id,
                    "name": device.name,
                    "user_agent": device.user_agent,
                    "last_ip_address": device.last_ip_address,
                    "login_count": device.login_count,
                    "first_seen_at": device.first_seen_at,
                    "last_seen_at": device.last_seen_at,
                    "is_active": device.id in active_device_ids,
                }
                for device in devices
            ],
        },
        "attempts": [
            {
                "id": attempt.id,
                "module_title": attempt.module.title,
                "module_type": attempt.module.module_type,
                "status": attempt.status,
                "started_at": attempt.started_at,
                "submitted_at": attempt.submitted_at,
                "graded_at": attempt.graded_at,
                "raw_score": str(attempt.raw_score) if attempt.raw_score is not None else None,
                "max_score": str(attempt.max_score) if attempt.max_score is not None else None,
                "graders": [
                    {
                        "id": grade.grader.id if grade.grader else None,
                        "name": (
                            f"{grade.grader.first_name} {grade.grader.last_name}"
                            if grade.grader
                            else "Pending"
                        ),
                        "email": grade.grader.email if grade.grader else None,
                        "part": grade.part.title,
                        "status": grade.status,
                        "graded_at": grade.graded_at,
                    }
                    for grade in attempt.part_grades
                ],
            }
            for attempt in attempts
        ],
    }


def dashboard_summary(db: Session, actor: User) -> dict:
    institute_id = _require_institute(actor)
    institute = db.get(Institute, institute_id)
    if institute is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Institute not found")

    members = list_members(db, actor)
    permissions = admin_permissions(actor)
    can_see_students = any(
        permissions[key]
        for key in (
            "view_students",
            "manage_students",
            "view_student_activity",
            "manage_student_sessions",
        )
    )
    visible_members = [
        member
        for member in members
        if (member["role"] == STUDENT and can_see_students)
        or (member["role"] == INST_INSTRUCTOR and permissions["manage_staff"])
    ]
    subscription = (
        subscription_service.subscription_status(db, institute_id)
        if permissions["view_billing"]
        else None
    )
    assigned_courses = (
        db.query(ExamModule)
        .join(InstituteModule, InstituteModule.module_id == ExamModule.id)
        .filter(
            InstituteModule.institute_id == institute_id,
            InstituteModule.is_active.is_(True),
            ExamModule.deleted_at.is_(None),
        )
        .order_by(InstituteModule.assigned_at.desc(), ExamModule.id.desc())
        .limit(6)
        .all()
    )
    return {
        "institute": {
            "id": institute.id,
            "name": institute.name,
            "slug": institute.slug,
            "contact_email": institute.contact_email,
            "is_active": institute.is_active,
        },
        "counts": {
            "students": sum(1 for member in members if member["role"] == STUDENT and not member["deleted_at"]),
            "instructors": sum(1 for member in members if member["role"] == INST_INSTRUCTOR and not member["deleted_at"]),
            "active_members": sum(1 for member in visible_members if member["is_active"]),
        },
        "subscription": subscription,
        # not gated on view_billing: expiry disables every account here, so the
        # countdown is shown to any institute admin who can open the dashboard
        "access": subscription_service.access_window(db, institute_id),
        "permissions": permissions,
        "recent_members": visible_members[:6],
        "assigned_courses": [
            {
                "id": module.id,
                "title": module.title,
                "slug": f"module-{module.id}",
                "summary": module.description,
                "level": module.module_type,
                "estimated_duration_minutes": module.duration_minutes,
            }
            for module in assigned_courses
        ],
    }
